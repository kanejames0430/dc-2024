
# Imports
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import networkx as nx
from bokeh.models import Range1d, Circle, MultiLine, ColorBar, LinearColorMapper, LogColorMapper
from bokeh.models.layouts import TabPanel, Tabs
from bokeh.plotting import figure
from bokeh.plotting import from_networkx
from bokeh.transform import linear_cmap, log_cmap
from bokeh.palettes import Inferno256
import ast
import math



def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total: 
        print()

def build_plots(graphs: dict[str, nx.Graph]) -> None:
    '''
    build plots of network x graphs
    Tiles them three to a row in plt. subplots
    pass any number
    '''
    
    rows = math.ceil(len(graphs)/3)
    cols = len(graphs)
    fig, ax = plt.subplots(rows, cols, figsize=(21, (7*rows)))
    for i, key in enumerate(graphs.keys()):
        nx.draw_networkx(graphs[key], ax=ax[i], with_labels=False, node_size=5)
        ax[i].set_title(key)
    return None

def build_graphs(datasets: dict[str, dict]) -> dict[str, nx.Graph]:
    '''
    build networkx graphs for all datasets in a dictionary
    of datasets returns them as a dictionary
    '''
    graph_dict = {}
    for key in datasets.keys():
        if type(datasets[key]) != list:
            graph_dict[key] = nx.Graph((datasets[key]))
    return graph_dict

def import_text(filename: str) -> dict[str, dict]:
    '''
    read dataset from textfiles, and return data as a dictionary
    in the following format
    {device: [list of contacts]}
    '''
    with open(filename, encoding='utf-8') as f:
        data = f.read()
    return (ast.literal_eval(data))

def trim_keys(data_dict: dict) -> dict:
    '''
    cut the key down to a reasonable label
    for each device in dataset works for filenames
    in dataset we used, may need adjustments for new filenaming
    conventions or new device types
    '''
    new_keys = {}
    for key in data_dict.keys():
        if 'Phone' in key:
            #start of new label
            s = key.rfind('Phone') + 8
            #end of new key
            e = key.rfind('/')
            new_keys[key[s:e]] = data_dict[key]
        elif 'SIM' in key:
            s = key.rfind('SIM')
            e = key.rfind('/')
            new_keys[key[s:e]] = data_dict[key]
        elif 'Tablet':
            s = key.rfind('Tablet') + 9
            e = key.rfind('/')
            new_keys[key[s:e]] = data_dict[key]
        else:
            print('*********')
            print('new type')
            print(key)
            print('*********')
    return new_keys


def component_graphs(graph, components_selection: int) -> dict[str, nx.Graph]:
    '''
    creates a dictionary containing the number of connected components
    specified in componenet_selection in descending order based on number of nodes
    '''
    core = sorted(nx.connected_components(graph),
                  key=len,
                  reverse=True)
    component_dict = {}
    if components_selection > len(core)+1:
        length = len(core)
        print(f'Graph only contains {length} components')
    else:
        for i in range(components_selection):
            component_dict[f'component {i+1}'] = graph.subgraph(core[i])
    return component_dict

def tabbed_plot(dictOfHash, graph_dict: dict[str, nx.Graph],
                node_highlights: list = []) -> Tabs:
    '''
    create a tabbed bokeh plot with tabs for all graphs in
    a graph dict.  utilizes keys to title tabs.
    '''
    figs_dict = {}
    panel_list = []
    for key, graph in graph_dict.items():
        figs_dict[key] = plot_graph_bokeh(dictOfHash, graph,
                                          title=key,
                                          node_highlights=node_highlights)
        panel_list.append(TabPanel(child=figs_dict[key], title=key))
    tabs = Tabs(tabs=panel_list)
    return tabs


def plot_graph_bokeh(dictOfHash, graph: nx.Graph,
                     title: str,
                     node_highlights: list = [],
                     file_color: str = 'firebrick' ,  # File nodes
                     hash_color: str = 'dodgerblue',  # Hash value nodes
                     dev_color: str = 'greenyellow'   # Device nodes
                     ) -> figure:
    '''
    plot a networkx graph in bokeh, creating hover labels, and
    highlighting nodes named in node_highlights list
    '''
    # set node color attributes in nx graph
    node_attrs = {}

    for node in graph.nodes():
        '''
        We are working with 3 kinda of nodes:
        1) if the node (checked as a string) is a hash value:
            then it exists as a value in the MD5 dictionary. so check there
            set the node color equal to respective hash color

        2) if the node has a json file name:
            then it exists as a key in a dictionary
            set the node color equal to respective device color

        3) else:
            it is the name of a file

        NOTE: The 'file type' to MD5 dictionary is smaller and easier to parse through
            AND contains all relevant hashes
        '''
        node_color = file_color
        if node in str(list(dictOfHash.values())):
            node_color = hash_color
        elif node in list(dictOfHash.keys()):
            node_color = dev_color

        node_attrs[node] = node_color
    nx.set_node_attributes(graph, node_attrs, 'node_color')
    # create bokeh plot
    HOVER_TOOLTIPS = [("Node", "@index")]
    plot = figure(tooltips=HOVER_TOOLTIPS,
                  tools='pan,wheel_zoom,save,reset',
                  active_scroll='wheel_zoom',
                  x_range=Range1d(-10.1, 10.1),
                  y_range=Range1d(-10.1, 10.1),
                  title=title)
    network_graph = from_networkx(graph,
                                  nx.spring_layout,
                                  scale=10,
                                  center=(0, 0))
    network_graph.node_renderer.glyph = Circle(radius=.75, fill_color='node_color')
    network_graph.edge_renderer.glyph = MultiLine(line_alpha=0.5, line_width=1)
    plot.renderers.append(network_graph)
    return plot


def cleanList(x) -> list:
    clean = []
    for item in x:
        if item != 'N/A':
            clean.append(item)
    return clean


def unique_phones(data_dictionary):
    """
    compares all phone numbers in data set to find close matches 
    in order to correct for missing country codes/prefixes
    """
    contact_list = []
    phonelookup = {}
    for phone, contacts in data_dictionary.items():
        contact_list.extend(contacts)
    contact_set = set(contact_list)
    for i in range(len(contact_set)):
        for j in range(i+1, len(contact_set)):
            cont1 = list(contact_set)[i]
            cont2 = list(contact_set)[j]
            short1 = cont1[-7:]
            short2 = cont2[-7:]
            if short1 == short2:
                if len(cont1) > len(cont2):
                    phonelookup[cont2] = cont1
                else:
                    phonelookup[cont1] = cont2
    for phone, contacts in data_dictionary.items():
        for i, contact in enumerate(contacts):
            if contact in phonelookup.keys():
                print("updated: ", data_dictionary[phone][i], " ", phonelookup[contact])
                data_dictionary[phone][i] = phonelookup[contact]
        data_dictionary[phone] = contacts
        # print(contacts_data[phone])
        # print(phonelookup)
    print(phonelookup)
    return data_dictionary


def merge_datasets(dataset1: dict, dataset2: dict) -> dict:
    '''
    merge two dict datasets
    '''
    merge_data = {}
    key_list = list(dataset1.keys()) + list(dataset2.keys())
    for key in key_list:
        if key in set(list(dataset1.keys())) and key in set(list(dataset2.keys())):
            merge_data[key] = dataset1[key] + dataset2[key]
        elif key in set(list(dataset1.keys())):
            merge_data[key] = dataset1[key]
        else:
            merge_data[key] = dataset2[key]
    return merge_data

def count_common_contacts(dataset: dict[str, list],
                          nodelist: list[str]) -> pd.DataFrame:
    '''
    counts number of common contacts in a dataset between nodes in nodelist.
    returns a pandas dataframe
    '''
    common_contacts = {}
    for i in range(len(nodelist)):
        common_contacts[nodelist[i]] = {}
        for j in range(i+1, len(nodelist)):
            shared_contacts = 0
            for contact in set(dataset[nodelist[i]]):
                if contact in dataset[nodelist[j]]:
                    shared_contacts += 1
                common_contacts[nodelist[i]][nodelist[j]] = shared_contacts
    cc_to_df = {'A': [], 'B': [], 'CommonContacts': []}
    for key1 in common_contacts.keys():
        for key2, value in common_contacts[key1].items():
            if value > 0:
                cc_to_df['A'].append(key1)
                cc_to_df['B'].append(key2)
                cc_to_df['CommonContacts'].append(value)
    cc_df = pd.DataFrame.from_dict(cc_to_df)
    return cc_df


def plot_colors_byscore(graph: nx.Graph, 
                        scores_dict: dict[str, dict],
                        color_mapping: str = 'linear') -> Tabs:
    '''
    create bokeh plots of graphs with nodes colored based on their scores.
    pass scores in a dictionary with scores for multiple metrics to get a 
    tabbed plot with a colored graph for each metric
    '''
    node_attrs = {}
    figs_dict = {}
    panel_list = []
    for key in scores_dict.keys():
        for key2, value in scores_dict[key].items():
            node_attrs[key2] = value
        nx.set_node_attributes(graph, node_attrs, key)
        HOVER_TOOLTIPS = [('Node', '@index'), ('Score', f'@{key}')]
        plot = figure(tooltips=HOVER_TOOLTIPS,
                      tools='pan,wheel_zoom,save,reset',
                      active_scroll='wheel_zoom',
                      x_range=Range1d(-10.1, 10.1),
                      y_range=Range1d(-10.1, 10.1))
        network_graph = from_networkx(graph,
                                      nx.spring_layout,
                                      scale=10,
                                      center=(0, 0))
        values = list(scores_dict[key].values())
        if color_mapping == 'log':
            cmap = log_cmap(
                field_name=key,
                palette=Inferno256,
                low=min(values),
                high=max(values))
            color_bar = ColorBar(color_mapper=LogColorMapper(
                palette=Inferno256,
                low=min(values),
                high=max(values)),
                label_standoff=12)
        else:
            cmap = linear_cmap(
                field_name=key,
                palette=Inferno256, low=min(values),
                high=max(values))
            color_bar = ColorBar(color_mapper=LinearColorMapper(
                palette=Inferno256,
                low=min(values),
                high=max(values)),
                label_standoff=12)
        network_graph.node_renderer.glyph = Circle(
            radius=10,
            fill_color=cmap)
        network_graph.edge_renderer.glyph = MultiLine(
            line_alpha=0.5,
            line_width=1)
        plot.renderers.append(network_graph)
        plot.add_layout(color_bar, 'right')
        figs_dict[key] = plot
        panel_list.append(
            TabPanel(child=figs_dict[key],
                  title=key))
    tabs = Tabs(tabs=panel_list)
    return tabs


def get_sheetnames_xlsx(filepath):
    '''
    gets the sheet names contained in an excel file given the filepath of an excel file
    '''
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


def sortDict(x):
    '''
    Sorts the dictionary by its values
    '''
    return {k: v for k, v in sorted(x.items(), key=lambda item: item[1])}



def plotVariablesByFreq(x):
    '''
    x is a dictionary. Draws a plot of the sheetnames (or variables) by
    frequency in the data
    '''
    xpoints = []
    ypoints = []


    freqs = {}

    #x is a dictionary, where (in this case) the keys are the file names, and the value is the list of sheetnames
    #create a new dictionary of sheetnames, with values as their value
    for item in x:
        for sheetname in list(x.get(item)):
            if sheetname in freqs:
                freqs[sheetname] += 1
            else: 
                freqs[sheetname] = 1
    
    #sort the dicionary so the graph looks nice
    freqs = sortDict(freqs)
        
    for item in freqs:
        xpoints.append(item)
        ypoints.append(freqs.get(item))


    #plot the graph
    plt.figure(figsize=(10, 6))  # Adjust the figure size as needed
    plt.bar(xpoints, ypoints)
    plt.xticks(rotation=45, ha='right', fontsize=8)  # Rotate x-axis labels
    plt.xlabel("Types of data contained in the dataset")
    plt.ylabel("Frequency of appearance")
    plt.tight_layout()  # Adjust layout to prevent clipping of labels
    plt.show()
